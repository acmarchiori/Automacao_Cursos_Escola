import unicodedata
import pandas as pd
import os
from sqlalchemy import create_engine
from dotenv import load_dotenv
from utils import processar_cursos, inserir_cursos
from openpyxl import load_workbook


load_dotenv()


# Configuração de conexão com o banco de dados SQL Server
def criar_engine():
    server = os.getenv("DB_SERVER")
    database = os.getenv("DB_DATABASE")
    username = os.getenv("DB_USERNAME")
    password = os.getenv("DB_PASSWORD")
    driver = os.getenv("DB_DRIVER")
    connection_string = (
        f"mssql+pyodbc://{username}:{password}@{server}:1433/{database}?"
        f"driver={driver}&Encrypt=yes&TrustServerCertificate=yes&Connection "
        f"Timeout=30"
    )
    return create_engine(connection_string)


# Normalizar o tipo de aula
def normalizar_tipo_aula(tipo_aula):
    if tipo_aula is None:
        return tipo_aula
    tipo_aula = tipo_aula.strip().upper()
    tipo_aula = unicodedata.normalize('NFKD', tipo_aula).encode(
        'ASCII', 'ignore').decode('utf-8')

    if "UMA RESPOSTA CORRETA" in tipo_aula or \
            "1 RESPOSTA CORRETA" in tipo_aula:
        return "ATIVIDADE MÚLTIPLA ESCOLHA COM 1 RESPOSTA CORRETA"
    if "N RESPOSTAS CORRETAS" in tipo_aula or \
            "N RESPOSTA CORRETA" in tipo_aula or \
            "N RESPOSTAS CORRETA" in tipo_aula:
        return "ATIVIDADE MÚLTIPLA ESCOLHA COM N RESPOSTAS CORRETAS"
    if "OPCOES CORRESPONDENTES" in tipo_aula or \
            "OPÇÕES CORRESPONDENTES" in tipo_aula:
        return "ATIVIDADE OPÇÕES CORRESPONDENTES"
    if "RESPOSTA ABERTA" in tipo_aula:
        return "ATIVIDADE COM RESPOSTA ABERTA"
    if "TEXTO E IMAGENS" in tipo_aula:
        return "CONTEÚDO EM TEXTO E IMAGENS"
    if "VIDEO" in tipo_aula or "VÍDEO" in tipo_aula:
        return "CONTEÚDO EM VÍDEO"
    if "ATIVIDADE" not in tipo_aula:
        if "MULTIPLA ESCOLHA COM 1 RESPOSTA CORRETA" in tipo_aula:
            return "ATIVIDADE MÚLTIPLA ESCOLHA COM 1 RESPOSTA CORRETA"
        if "MULTIPLA ESCOLHA COM N RESPOSTAS CORRETAS" in tipo_aula or \
            "MULTIPLA ESCOLHA COM N RESPOSTA CORRETA" in tipo_aula or \
                "MULTIPLA ESCOLHA COM N RESPOSTAS CORRETA" in tipo_aula:
            return "ATIVIDADE MÚLTIPLA ESCOLHA COM N RESPOSTAS CORRETAS"
        if "OPCOES CORRESPONDENTES" in tipo_aula or \
                "OPÇÕES CORRESPONDENTES" in tipo_aula:
            return "ATIVIDADE OPÇÕES CORRESPONDENTES"
        if "RESPOSTA ABERTA" in tipo_aula:
            return "ATIVIDADE COM RESPOSTA ABERTA"
    return tipo_aula


def carregar_dados():
    # Caminho para a pasta contendo as planilhas
    pasta_path = "/home/acmarchiori/Área de Trabalho/Importações Mosaico/"

    # Listar todos os arquivos na pasta
    arquivos = [f for f in os.listdir(pasta_path) if f.endswith('.xlsm')]

    # Iterar sobre os arquivos e carregar os dados de cada planilha
    for arquivo in arquivos:
        file_path = os.path.join(pasta_path, arquivo)
        print(f"Carregando planilha: {file_path}")

        wb = load_workbook(filename=file_path, data_only=True)
        ws = wb["Plan1"]

        # Extrair os dados da planilha
        data = ws.values
        columns = next(data)[0:]
        df = pd.DataFrame(data, columns=columns)

        # Imprimir as colunas do DataFrame
        print("Colunas disponíveis no DataFrame:", df.columns.tolist())

        # Renomear colunas conforme necessário
        df.rename(
            columns={
                "ANO": "ANO_ESCOLAR",
                "SEGMENTO": "SEGMENTO_ESCOLAR",
                "CURSO": "TITULO",
                "MÓDULO": "NOME_MODULO",
                "ORDEM MÓDULO": "ORDEM_MODULO",
                "CAPÍTULO": "NOME_CAPITULO",
                "ORDEM CAPÍTULO": "ORDEM_CAPITULO",
                "AULA": "TITULO_AULA",
                "ORDEM AULA": "ORDEM_AULA",
                "BNCC": "CODIGOS_BNCC",
                "PALAVRAS CHAVES": "PALAVRAS_CHAVES",
                "NÍVEL": "NIVEL",
                "LINK/CONTEÚDO": "LINK_CONTEUDO",
                "ESTRATÉGIA DE APRENDIZAGEM": "ESTRATEGIA_APRENDIZAGEM",
                "TIPO DE AULA": "TIPO_AULA",
            },
            inplace=True,
        )

    # Normalizar valores de ANO_ESCOLAR para garantir correspondência
    # com o banco de dados
    df['ANO_ESCOLAR'] = df['ANO_ESCOLAR'].str.strip().str.replace('º', '°')

    # Garantir que valores nulos ou vazios sejam convertidos para None
    df = df.where(pd.notnull(df), None)

    # Normalizar valores de TIPO_AULA para garantir correspondência
    # com o banco de dados
    df['TIPO_AULA'] = df['TIPO_AULA'].apply(normalizar_tipo_aula)

    # Converter valores de NIVEL para inteiros e formatar corretamente
    df['NIVEL'] = df[
        'NIVEL'].apply(lambda x: f"Nível {int(x)}" if pd.notnull(x) else None)

    # Verificar se a coluna ESTRATEGIA_APRENDIZAGEM
    # está sendo carregada corretamente
    print("Valores da coluna ESTRATEGIA_APRENDIZAGEM:",
          df['ESTRATEGIA_APRENDIZAGEM'].unique())

    # Encontrar o índice da coluna "LINK_CONTEÚDO"
    link_col_idx = None
    for idx, cell in enumerate(ws[1]):
        if cell.value == "LINK/CONTEÚDO":
            link_col_idx = idx + 1
            break

    if link_col_idx is None:
        raise ValueError("Coluna 'LINK/CONTEÚDO' não encontrada na planilha.")

# Extrair links do Google Drive
    for row in ws.iter_rows(
        min_row=2, max_row=ws.max_row,
        min_col=link_col_idx,
        max_col=link_col_idx
    ):
        for cell in row:
            if cell.hyperlink:
                df.at[cell.row - 2, 'LINK_CONTEUDO'] = cell.hyperlink.target

    # Processar os dados
    cursos_df = processar_cursos(df)

    # Criar engine de conexão
    engine = criar_engine()

    # Chamar a função para inserir os cursos
    inserir_cursos(engine, cursos_df)
    print("Processamento completo. Todas as linhas foram processadas.")


if __name__ == "__main__":
    carregar_dados()
